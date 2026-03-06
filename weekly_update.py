#!/usr/bin/env python3
"""
KRANK Weekly Update Script
──────────────────────────
1. Drop KRANK_Social_Tracker.xlsx into this folder
2. Run:  python3 weekly_update.py
3. Script updates index.html, commits, and pushes automatically
"""

import re
import subprocess
from datetime import date

import openpyxl

# ── CONFIG (mirrors krank.config) ─────────────────────────
XLSX_FILE  = 'KRANK_Social_Tracker.xlsx'
SHEET_NAME = 'Weekly Tracker'
HTML_FILE  = 'index.html'

COL_NAME      = 2   # Column B  – celebrity name
COL_FOLLOWERS = 28  # Column AB – Latest Followers (in millions)
COL_FAN       = 30  # Column AD – Latest Fan Score (0–100)
# ──────────────────────────────────────────────────────────


def col_idx(letter):
    """Excel column letter(s) to 1-based index."""
    n = 0
    for c in letter.upper():
        n = n * 26 + (ord(c) - ord('A') + 1)
    return n


def load_sheet():
    print(f"Reading {XLSX_FILE}  /  sheet: {SHEET_NAME} ...")
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(
            f"ERROR: sheet '{SHEET_NAME}' not found.\n"
            f"Available sheets: {wb.sheetnames}"
        )
    return wb[SHEET_NAME]


def build_update_map(ws):
    """Return {NAME_UPPER: (followers, fan_score)} from the sheet."""
    updates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name      = row[COL_NAME - 1]
        followers = row[COL_FOLLOWERS - 1]
        fan_score = row[COL_FAN - 1]

        if name is None:
            continue
        key = str(name).strip().upper()
        if followers is not None or fan_score is not None:
            updates[key] = (followers, fan_score)

    print(f"Found {len(updates)} celebrities in spreadsheet.")
    return updates


def update_html(updates):
    with open(HTML_FILE, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    matched = []
    skipped = []

    for i, line in enumerate(lines):
        # Each DB entry is a single line containing  name:'...'
        m = re.search(r"name:'([^']+)'", line)
        if not m:
            continue

        name_in_html = m.group(1).strip().upper()
        if name_in_html not in updates:
            continue

        followers, fan_score = updates[name_in_html]
        original = line

        # ── followers ──────────────────────────────────────
        if followers is not None:
            try:
                fval = float(followers)
                # Keep one decimal place; strip trailing .0 if whole number
                fstr = f"{fval:.1f}" if fval != int(fval) else str(int(fval))
                line = re.sub(r"followers:[^,}]+", f"followers:{fstr}", line)
            except (ValueError, TypeError):
                pass  # leave unchanged if value is unexpected

        # ── fanActivity ────────────────────────────────────
        if fan_score is not None:
            try:
                fval = int(round(float(fan_score)))
                fval = max(0, min(100, fval))  # clamp to 0–100
                line = re.sub(r"fanActivity:\d+", f"fanActivity:{fval}", line)
            except (ValueError, TypeError):
                pass

        if line != original:
            lines[i] = line
            matched.append(m.group(1))

    with open(HTML_FILE, 'w', encoding='utf-8') as f:
        f.writelines(lines)

    # Report celebrities in sheet but not found in HTML
    html_names = set()
    for line in lines:
        m = re.search(r"name:'([^']+)'", line)
        if m:
            html_names.add(m.group(1).strip().upper())
    for key in updates:
        if key not in html_names:
            skipped.append(key)

    print(f"Updated : {len(matched)} celebrities")
    if matched:
        for n in matched:
            print(f"  ✓ {n}")
    if skipped:
        print(f"No match: {len(skipped)} (check spelling vs index.html)")
        for n in skipped:
            print(f"  ? {n}")


def git_commit_push():
    today = date.today().strftime('%Y-%m-%d')
    msg   = f"Weekly update – {today}"

    subprocess.run(['git', 'add', HTML_FILE], check=True)

    # Check if there is anything to commit
    result = subprocess.run(
        ['git', 'diff', '--cached', '--quiet'],
        capture_output=True
    )
    if result.returncode == 0:
        print("No changes detected – nothing to commit.")
        return

    subprocess.run(['git', 'commit', '-m', msg], check=True)
    subprocess.run(['git', 'push'], check=True)
    print(f"\n✓ Pushed: \"{msg}\"")


def main():
    ws      = load_sheet()
    updates = build_update_map(ws)
    update_html(updates)
    git_commit_push()
    print("\nDone.")


if __name__ == '__main__':
    main()
